from ast import Num
from operator import index
from re import sub
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import time

# Pending...

#book = Workbook()
#sheet = book.active
book = Workbook()
sheet = book.active

#def makesheet():
 #   sheet["B1"] = "Hola mundo"
  #  sheet["A2"] = "Pato"
    #sheet["E3"] = plA

#makesheet()
num = 5
sub = 1

for i in range(num) :
    time.sleep(1)
    sheet[f"B{sub}"] = "Mensaje.ex-cel"
    book.save("C:/Users/franc/Escritorio/BotDiscPython/DataBase/pruebaE.xlsx")
    print(sub)
    sub = sub + 1


